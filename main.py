# This is a sample Python script.
import openpyxl
from openpyxl import load_workbook
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    workbook = load_workbook(filename="Book4.xlsx", data_only=True)
    sheet1 = workbook["Nodes"]
    sheet2 = workbook["Edges"]
    sheet3 = workbook["edgelist"]

    #for i in range(2, sheet1.max_row+1):
        #sheet3.cell(i, 2).value = sheet1.cell(i, 3).value
        #sheet3.cell(i, 3).value = sheet1.cell(i, 4).value

    for i in range(2, sheet2.max_row+1):
        temp = int(sheet2.cell(i, 4).value)
        for j in range(2, sheet1.max_row+1):
            if temp == int(sheet1.cell(j, 2).value):
                sheet2.cell(i, 7).value = sheet1.cell(j, 1).value

    workbook.save('Book5.xlsx')






# See PyCharm help at https://www.jetbrains.com/help/pycharm/
